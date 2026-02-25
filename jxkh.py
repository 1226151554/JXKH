from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file, abort, \
    make_response, send_from_directory
from login_code import generate_login_codes_by_role, export_login_codes
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from datetime import datetime
from database import db
from io import BytesIO
import pandas as pd
import yaml
import os
import re

# 初始化Flask应用
app = Flask(__name__)

# 上传配置
UPLOAD_ROOT = os.path.join(app.root_path, 'uploads', 'zdgz')
ALLOWED_EXT = {'pdf', 'zip', 'xls', 'xlsx'}


def allowed_file(filename):
    """
    检查文件扩展名是否在允许的范围内

    Args:
        filename: 文件名

    Returns:
        bool: 文件扩展名是否允许
    """
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT


# 加载配置
with open('config.yaml', 'r') as f:
    config = yaml.safe_load(f)
    app.config['SECRET_KEY'] = config['app']['secret_key']


# ==================== 前台用户路由 ====================

@app.route('/')
def index():
    """
    首页路由 - 显示重点工作指标和满意度部门信息

    功能:
    - 检查用户登录状态
    - 获取当前用户角色权限下的重点工作指标
    - 获取当前用户角色权限下的满意度部门
    - 渲染首页模板
    """
    # 检查用户是否已登录
    if 'login_code' not in session:
        return redirect(url_for('login'))

    role_id = session.get('role_id')

    # ===== 重点工作指标 =====
    zdgz_list = db.get_zdgz()

    # 获取角色对应的重点工作指标权限（部门名）
    role_zdgz_perm = db.get_role_zdgz_permissions()
    allowed_zdgz_depts = set(role_zdgz_perm.get(role_id, []))

    # 过滤重点工作指标，只保留当前角色有权限的部门
    zdgz_by_dept = {}
    for item in zdgz_list:
        dept = item['department']
        if dept in allowed_zdgz_depts:
            zdgz_by_dept.setdefault(dept, []).append(item)

    # ===== 满意度部门 =====
    all_departments = db.get_departments()
    myd_perm_map = db.get_myd_permissions()
    allowed_depts = set(myd_perm_map.get(role_id, {}).keys())

    # 只保留当前角色有权限的部门
    departments = [
        d for d in all_departments
        if d['id'] in allowed_depts
    ]

    return render_template(
        'index.html',
        zdgz_by_dept=zdgz_by_dept,
        departments=departments
    )


@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    登录路由 - 处理用户登录

    功能:
    - GET: 显示登录页面
    - POST: 验证登录码和密码
    """
    if request.method == 'POST':
        login_code = request.form['login_code'].strip()
        password = request.form['password']
        ip = request.remote_addr

        user = db.yz_user(login_code)

        if not user:
            return render_template('login.html', error="无效的登录码")

        if user['used'] == 1:
            return render_template('login.html', error="该登录码已使用过")

        if user['password'] == password:
            session['login_code'] = login_code
            session['role_id'] = user['role_id']
            db.login_rec(ip, login_code)
            return redirect(url_for('index'))

    return render_template('login.html')


@app.route('/score/save', methods=['POST'])
def save_score():
    """
    保存评分路由 - 保存用户评分结果

    功能:
    - 验证用户登录状态
    - 后端校验重点工作指标优秀率（≤60%）
    - 保存重点工作指标评分
    - 保存满意度评分
    - 标记登录码为已使用
    """
    role_id = session.get('role_id')
    login_code = session.get('login_code')

    if not role_id or not login_code:
        return redirect(url_for('login'))

    # ========= 后端兜底校验（重点工作指标优秀率 ≤ 60%）=========
    zdgz_scores = []

    for key, value in request.form.items():
        if key.startswith('zdgz_'):
            try:
                score = float(value)
            except ValueError:
                score = 0
            zdgz_scores.append(score)

    total_cnt = len(zdgz_scores)
    if total_cnt > 0:
        max_excellent = int(total_cnt * 0.6)
        excellent_cnt = sum(1 for s in zdgz_scores if s >= 120)

        if excellent_cnt > max_excellent:
            return (
                f"重点工作指标中，“优秀”最多只能选择 {max_excellent} 项（60%），"
                f"当前选择了 {excellent_cnt} 项，请返回重新评分。",
                400
            )
    # ========= 校验结束 =========

    # 重点工作指标评分
    for key, value in request.form.items():
        if key.startswith('zdgz_'):
            zdgz_id = int(key.replace('zdgz_', ''))
            score = float(value)

            db.save_zdgz_score(
                login_code=login_code,
                role_id=role_id,
                zdgz_id=zdgz_id,
                score=score
            )

    # 满意度评分
    for key, value in request.form.items():
        if key.startswith('satisfaction_'):
            dept_id = int(key.replace('satisfaction_', ''))
            score = float(value)

            db.save_myd_score(
                login_code=login_code,
                role_id=role_id,
                dept_id=dept_id,
                score=score
            )

    # 标记登录码已使用
    db.set_used(login_code)
    session.clear()

    return render_template('score_success.html')


@app.route('/logout')
def logout():
    """
    登出路由 - 清除用户会话

    功能:
    - 清除登录码会话信息
    - 重定向到登录页
    """
    session.pop('login_code', None)
    return redirect(url_for('login'))


# ==================== 管理员路由 ====================

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """
    管理员登录路由

    功能:
    - GET: 显示管理员登录页面
    - POST: 验证管理员用户名和密码
    """
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        admin = db.admin_login(username, password)
        if admin:
            session['admin_user'] = admin['username']
            return redirect(url_for('admin_dashboard'))
        else:
            return render_template('admin/login.html', error='用户名或密码错误')

    return render_template('admin/login.html')


def admin_required(func):
    """
    管理员权限装饰器

    功能:
    - 检查用户是否已登录为管理员
    - 未登录则重定向到管理员登录页
    """

    def wrapper(*args, **kwargs):
        if 'admin_user' not in session:
            return redirect(url_for('admin_login'))
        return func(*args, **kwargs)

    wrapper.__name__ = func.__name__
    return wrapper


@app.route('/admin/index')
@admin_required
def admin_dashboard():
    """
    管理后台首页路由

    功能:
    - 需要管理员权限
    - 渲染管理后台首页
    """
    return render_template('admin/index.html')


@app.route('/admin/myd', methods=['GET', 'POST'])
def satisfaction_manage():
    """
    满意度相关信息维护页面

    功能:
    - GET: 显示满意度部门管理页面
    - POST: 添加新部门
    """
    if request.method == 'POST':
        dept_name = request.form.get('dept_name')
        dept_type = request.form.get('dept_type')
        if dept_name:
            db.add_department(dept_name, dept_type)

    departments = db.get_departments()
    roles = db.get_roles()

    myd_permissions = db.get_myd_permissions()

    return render_template(
        '/admin/myd.html',
        departments=departments,
        roles=roles,
        myd_permissions=myd_permissions
    )


@app.route('/admin/myd/permission/save', methods=['POST'])
def save_myd_permission():
    """
    保存满意度配置：
    - 角色-部门权限
    - 角色-部门满意度权重
    """
    try:
        data = request.json
        if not isinstance(data, list):
            return jsonify({'error': '数据格式错误'}), 400

        db.save_myd_permissions(data)

        return jsonify({'msg': '满意度权限与权重保存成功'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/admin/role/add', methods=['POST'])
def add_role():
    """
    新增角色路由

    功能:
    - 验证角色名只能为中文
    - 创建新角色
    """
    role_name = request.json['role_name']

    if not all('\u4e00' <= c <= '\u9fff' for c in role_name):
        return jsonify({'error': '只能输入中文'}), 400

    try:
        db.create_role(role_name)
    except:
        return jsonify({'error': '角色名已存在'}), 400

    return jsonify({'status': 'ok'})


@app.route('/admin/role/delete', methods=['POST'])
def delete_role():
    """
    删除角色路由

    功能:
    - 删除角色-部门权限关系
    - 删除角色本身
    """
    role_id = int(request.json['role_id'])

    with db.get_connection() as conn:
        cursor = conn.cursor()

        # 删除角色-部门权限关系
        cursor.execute(
            "DELETE FROM role_dept_permission WHERE role_id=%s",
            (role_id,)
        )

        # 删除角色本身
        cursor.execute(
            "DELETE FROM evaluator_role WHERE id=%s",
            (role_id,)
        )

        conn.commit()

    return jsonify({'status': 'ok'})


@app.route('/admin/departments/delete/<int:dept_id>')
@admin_required
def delete_department(dept_id):
    """
    删除部门路由

    功能:
    - 需要管理员权限
    - 删除指定部门
    """
    db.delete_department(dept_id)
    return redirect(url_for('admin_departments'))


@app.route('/admin/departments/update_desc', methods=['POST'])
def update_dept_desc():
    """
    部门工作完成情况编辑路由

    功能:
    - 更新部门工作完成情况说明
    - 校验字数（300-1000字）
    """
    dept_id = request.form.get('dept_id')
    work_desc = request.form.get('work_desc', '').strip()

    # 允许为空，但如果填写了就校验字数
    if work_desc and not (300 <= len(work_desc) <= 1500):
        return "工作完成情况说明需控制在300-1000字之间", 400

    with db.get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE department SET work_desc=%s WHERE id=%s",
            (work_desc, dept_id)
        )
        conn.commit()

    return redirect('/admin/myd?saved=1')


# ==================== 重点工作指标管理路由 ====================

@app.route('/admin/zdgz')
def zdgz_page():
    """
    打分项管理路由

    功能:
    - 获取角色信息
    - 获取重点工作指标部门信息
    - 获取角色重点工作指标权限
    - 获取重点工作指标项目
    - 渲染重点工作指标管理页面
    """
    roles = db.get_roles()
    zdgz_departments = db.get_zdgz_departments()
    zdgz_permissions = db.get_role_zdgz_permissions()
    zdgz_items = db.get_zdgz()

    return render_template(
        '/admin/zdgz.html',
        roles=roles,
        zdgz_departments=zdgz_departments,
        zdgz_permissions=zdgz_permissions,
        zdgz_items=zdgz_items
    )


@app.route('/admin/zdgz/import', methods=['POST'])
def import_zdgz():
    """
    上传重点工作指标信息路由

    功能:
    - 从Excel文件导入重点工作指标数据
    - Excel 第一行是表头，数据从 A2 开始
    - A: 部门
    - B: 绩效指标
    - C: 指标含义
    - D: 完成情况（300-1000字）
    """
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '未选择文件'}), 400

    try:
        wb = load_workbook(file)
        sheet = wb.active

        departments = []
        indicators = []
        descriptions = []
        work_descs = []

        current_department = None

        # 从第 2 行开始读取（跳过表头）
        for row in range(2, sheet.max_row + 1):
            department = sheet[f'A{row}'].value
            if department:
                current_department = str(department).strip()
            else:
                department = current_department

            indicator_name = sheet[f'B{row}'].value
            description = sheet[f'C{row}'].value
            work_desc = sheet[f'D{row}'].value

            # B、C、D 都为空，认为到末尾
            if not indicator_name and not description and not work_desc:
                break

            # 基础清洗
            indicator_name = db.clean_text(indicator_name) if indicator_name else None
            description = db.clean_text(description) if description else "无指标含义"
            work_desc = db.clean_text(work_desc)

            # 完成情况字数校验
            if work_desc:
                length = len(work_desc)
                if length < 1 or length > 1000:
                    return jsonify({
                        'error': f'第 {row} 行完成情况字数为 {length}，需在 1–1000 字之间'
                    }), 400

            departments.append(department)
            indicators.append(indicator_name)
            descriptions.append(description)
            work_descs.append(work_desc)

        data = pd.DataFrame({
            'department': departments,
            'indicator_name': indicators,
            'description': descriptions,
            'work_desc': work_descs
        })

        if data.empty:
            return jsonify({'error': 'Excel 中未读取到任何指标数据'}), 400

        insert_count = 0

        with db.get_connection() as conn:
            cursor = conn.cursor()

            # 清空旧数据
            cursor.execute("DELETE FROM zdgz")

            # 插入新数据
            for _, row in data.iterrows():
                cursor.execute("""
                    INSERT INTO zdgz (
                        department,
                        indicator_name,
                        description,
                        work_desc
                    )
                    VALUES (%s, %s, %s, %s)
                """, (
                    row['department'],
                    row['indicator_name'],
                    row['description'],
                    row['work_desc']
                ))
                insert_count += 1

            conn.commit()

        return jsonify({'msg': f'导入成功，已更新 {insert_count} 条重点工作指标'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/admin/zdgz/evidence/upload', methods=['POST'])
@admin_required
def upload_zdgz_evidence():
    """
    上传佐证文件路由

    功能:
    - 需要管理员权限
    - 处理重点工作指标佐证文件上传
    - 安全处理文件名
    - 删除旧文件并保存新文件
    - 更新数据库中的文件路径
    """
    zdgz_id = request.form.get('zdgz_id')
    file = request.files.get('file')

    if not zdgz_id or not file:
        return jsonify({'error': '参数不完整'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': '不支持的文件类型'}), 400

    # ===== 查询指标名称 =====
    zdgz = db.get_zdgz_by_id(zdgz_id)
    if not zdgz:
        return jsonify({'error': '指标不存在'}), 404

    indicator_name = zdgz['indicator_name']

    # 文件名安全处理
    indicator_name = re.sub(r'[\\/:*?"<>|]', '_', indicator_name)

    # ===== 删除旧文件 =====
    old_path = db.get_zdgz_evidence_path(zdgz_id)
    if old_path:
        abs_old = os.path.join(app.root_path, old_path)
        if os.path.exists(abs_old):
            os.remove(abs_old)

    # ===== 保存新文件 =====
    ext = file.filename.rsplit('.', 1)[1].lower()
    save_dir = os.path.join(UPLOAD_ROOT, str(zdgz_id))
    os.makedirs(save_dir, exist_ok=True)

    filename = f"{indicator_name}佐证材料.{ext}"
    abs_path = os.path.join(save_dir, filename)
    file.save(abs_path)

    rel_path = f"uploads/zdgz/{zdgz_id}/{filename}"

    # ===== 更新数据库 =====
    db.update_zdgz_evidence(zdgz_id, rel_path)

    return jsonify({'msg': '上传成功'})


@app.route('/zdgz/evidence/<int:zdgz_id>')
def download_zdgz_evidence(zdgz_id):
    """
    下载佐证材料路由

    功能:
    - 根据指标ID下载对应的佐证材料
    """
    path = db.get_zdgz_evidence_path(zdgz_id)
    if not path:
        abort(404)

    # 拆目录和文件名
    directory = os.path.dirname(path)
    filename = os.path.basename(path)

    return send_from_directory(
        directory,
        filename,
        as_attachment=True
    )


@app.route('/admin/zdgz/permission/save', methods=['POST'])
def save_zdgz_permission():
    """
    保存重点工作指标：
    - 角色-部门权限
    - 角色在重点工作中的权重
    """
    try:
        data = request.json
        if not isinstance(data, list):
            return jsonify({'error': '数据格式错误'}), 400

        # 保存角色-部门权限
        db.save_role_zdgz_permissions(data)

        # 更新角色重点工作权重
        db.update_role_zdgz_weights(data)

        return jsonify({'msg': '重点工作指标权限与权重保存成功'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/admin/login_codes', methods=['GET', 'POST'])
@admin_required
def admin_login_codes():
    """
    管理员登录码路由

    功能:
    - GET: 显示登录码管理页面
    - POST: 生成新的登录码并导出Excel
    """
    roles = db.get_roles()

    if request.method == 'POST':

        # ===== 清空历史数据 =====
        db.clear_all_scores()  # 清空 zdgz_score / myd_score
        db.clear_login_codes()  # 清空旧登录码

        # ===== 生成新的登录码 =====
        role_count_map = {}
        for r in roles:
            count = int(request.form.get(f'role_{r["id"]}', 0))
            if count > 0:
                role_count_map[r['id']] = count

        generate_login_codes_by_role(role_count_map)

        # ===== 导出 Excel =====
        excel_io = export_login_codes()

        return send_file(
            excel_io,
            as_attachment=True,
            download_name=f"登录码_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    return render_template('/admin/login_codes.html', roles=roles)


@app.route('/admin/scores')
def admin_scores():
    """
    管理员查看评分结果页面路由

    功能:
    - 获取登录码统计信息
    - 获取重点工作指标评分汇总
    - 获取满意度评分汇总
    - 渲染评分结果页面
    """
    stats = db.get_login_code_stats_by_role()

    zdgz_scores = db.get_zdgz_score_summary()
    myd_scores = db.get_myd_score_summary()

    return render_template(
        'admin/scores.html',
        total_count=stats['total_count'],
        used_count=stats['used_count'],
        role_stats=stats['roles'],
        zdgz_scores=zdgz_scores,
        myd_scores=myd_scores
    )


@app.route('/admin/scores/export')
def export_scores():
    """
    导出评分结果路由

    功能:
    - 导出重点工作指标评分Excel
    - 导出满意度评分Excel
    - 合并为一个Excel文件下载
    """
    zdgz_df = db.export_zdgz_score_excel()
    myd_df = db.export_myd_score_excel()

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        zdgz_df.to_excel(writer, index=False, sheet_name='重点工作指标')
        myd_df.to_excel(writer, index=False, sheet_name='满意度评价')

    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='绩效考核评分汇总.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == '__main__':
    app.run(debug=True, host=config['app']['host'], port=config['app']['port'])